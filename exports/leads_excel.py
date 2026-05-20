"""Build an .xlsx workbook of CRM leads for the contextologist.

Marketing column set — no client names/phones (PII stays out). The
caller (export_leads_to_excel tool handler) hands in leads already
enriched by ToolHandlers.get_leads logic: each dict has the raw Bitrix
fields plus a resolved `direction` label and a `card_url`.
"""

from io import BytesIO
from typing import Any, Dict, List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


# Bitrix lead STATUS_ID → Russian label. Growzone uses the standard
# funnel; custom UC_* statuses fall back to STATUS_SEMANTIC_ID.
LEAD_STATUS_RU: Dict[str, str] = {
    "NEW": "Новый",
    "IN_PROCESS": "В работе",
    "PROCESSED": "Обработан",
    "CONVERTED": "Квал лид",
    "JUNK": "Неквал лид",
}

_SEMANTIC_RU: Dict[str, str] = {
    "P": "В работе",
    "S": "Квал лид",
    "F": "Неквал лид",
}

# (header, lead-dict key) — order defines column order.
_COLUMNS = [
    ("ID", "ID"),
    ("Дата создания", "_date"),
    ("Название лида", "TITLE"),
    ("Телефон", "_phone"),
    ("Статус", "_status"),
    ("Источник (код)", "SOURCE_ID"),
    ("Источник (описание)", "SOURCE_DESCRIPTION"),
    ("Направление", "direction"),
    ("UTM Source", "UTM_SOURCE"),
    ("UTM Medium", "UTM_MEDIUM"),
    ("UTM Campaign", "UTM_CAMPAIGN"),
    ("UTM Content", "UTM_CONTENT"),
    ("UTM Term", "UTM_TERM"),
    ("Причина отказа", "UF_CRM_1723465843"),
    ("Ссылка на карточку", "card_url"),
]

_HEADER_FILL = PatternFill("solid", fgColor="1F6F43")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_MAX_COL_WIDTH = 55


def _status_ru(lead: Dict[str, Any]) -> str:
    """Human-readable lead status, resilient to custom UC_* codes."""
    status_id = (lead.get("STATUS_ID") or "").strip()
    if status_id in LEAD_STATUS_RU:
        return LEAD_STATUS_RU[status_id]
    semantic = (lead.get("STATUS_SEMANTIC_ID") or "").strip()
    if semantic in _SEMANTIC_RU:
        return _SEMANTIC_RU[semantic]
    return status_id or "—"


def _phone(lead: Dict[str, Any]) -> str:
    """Phone number(s) from the Bitrix PHONE multifield.

    PHONE is a list of {'VALUE': '+7...', 'VALUE_TYPE': 'MOBILE', ...};
    several numbers are joined with '; '.
    """
    raw = lead.get("PHONE")
    if not raw:
        return ""
    if isinstance(raw, list):
        nums = [
            str(p.get("VALUE", "")).strip()
            for p in raw
            if isinstance(p, dict) and p.get("VALUE")
        ]
        return "; ".join(n for n in nums if n)
    return str(raw).strip()


def _fmt_date(raw: Any) -> str:
    """Bitrix DATE_CREATE '2026-05-19T10:47:14+03:00' → '2026-05-19 10:47'."""
    if not raw or not isinstance(raw, str):
        return ""
    s = raw.replace("T", " ")
    # Drop timezone suffix and seconds.
    s = s.split("+")[0].strip()
    parts = s.split(":")
    if len(parts) >= 2:
        return f"{parts[0]}:{parts[1]}"
    return s


def build_leads_xlsx(leads: List[Dict[str, Any]]) -> bytes:
    """Render leads into an .xlsx file, returned as raw bytes.

    Empty `leads` still produces a valid workbook with just the header
    row, so the caller can always send a file.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Лиды"

    headers = [h for h, _ in _COLUMNS]
    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for lead in leads:
        row = []
        for _, key in _COLUMNS:
            if key == "_date":
                row.append(_fmt_date(lead.get("DATE_CREATE")))
            elif key == "_status":
                row.append(_status_ru(lead))
            elif key == "_phone":
                row.append(_phone(lead))
            else:
                val = lead.get(key)
                row.append("" if val is None else str(val))
        ws.append(row)

    # Auto-ish column widths — sample the header + every row, cap the width.
    for col_idx, (header, _) in enumerate(_COLUMNS, start=1):
        longest = len(header)
        for row_idx in range(2, ws.max_row + 1):
            v = ws.cell(row=row_idx, column=col_idx).value
            if v is not None:
                longest = max(longest, len(str(v)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(
            longest + 2, _MAX_COL_WIDTH
        )

    ws.freeze_panes = "A2"
    if ws.max_row >= 1:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
